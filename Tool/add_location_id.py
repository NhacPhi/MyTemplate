import openpyxl

wb = openpyxl.load_workbook('D:/UnityUdeme/MyTemplate/MyTemplate/Tool/data/GameNarrative.xlsx')
ws = wb['Actors']

# Check if LocationID already exists
has_location_id = False
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == 'LocationID':
        has_location_id = True
        break

if not has_location_id:
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = 'LocationID'
    
    # Fill default value if you want, e.g., 'Location_Ocean'
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            # Setting default to a placeholder, or leave blank. Let's leave blank and let user fill it.
            pass

    wb.save('D:/UnityUdeme/MyTemplate/MyTemplate/Tool/data/GameNarrative.xlsx')
    print("Added LocationID column successfully.")
else:
    print("LocationID column already exists.")
