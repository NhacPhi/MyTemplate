import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

desc_vn = "Tăng {0}% Tấn Công và {0}% Phòng Thủ. Trạng thái Áp Đảo (HP > 50%) tăng 16% Sát thương Kỹ năng. Trạng thái Huyết Chiến (HP ≤ 50%) nhận 16% Hút Máu tương đương 16% tổng sát thương gây ra."
desc_en = "Increases ATK by {0}% and DEF by {0}%. In Overpower state (HP > 50%), increases Skill Damage by 16%. In Bloodbath state (HP ≤ 50%), gains 16% Lifesteal."

# 1. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR'] if 'STR' in wb_loc else wb_loc.active

# Check if key exists
key_found = False
for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_FERROUS_CHAOS_SKILL':
        row[1].value = desc_en
        row[2].value = desc_vn
        key_found = True
        break
if not key_found:
    ws_str.append(['STR_FERROUS_CHAOS_SKILL', desc_en, desc_vn])

wb_loc.save(loc_path)
print("Updated Localizations.xlsx")

# Update JSON localizations
for lang_file, text in [('Assets/Data/Localization/Localization_VIETNAMESE.json', desc_vn),
                       ('Assets/Data/Localization/Localization_ENGLISH.json', desc_en)]:
    if os.path.exists(lang_file):
        with open(lang_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        data['STR_FERROUS_CHAOS_SKILL'] = text
        with open(lang_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"Updated {lang_file}")

# 2. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# Passives Sheet
ws_p = wb_gc['Passives']
p_found = False
for row in ws_p.iter_rows(values_only=False):
    if row[0].value == 'psv_ferrous_chaos':
        row[1].value = 'STR_FERROUS_CHAOS_SKILL'
        row[2].value = desc_vn
        p_found = True
        break
if not p_found:
    ws_p.append(['psv_ferrous_chaos', 'STR_FERROUS_CHAOS_SKILL', desc_vn])

# StaticModifiers Sheet
ws_sm = wb_gc['StaticModifiers']
# Remove old psv_ferrous_chaos if any
rows_to_keep = []
for r in list(ws_sm.iter_rows(values_only=True)):
    if r[0] != 'psv_ferrous_chaos':
        rows_to_keep.append(r)
# Rebuild StaticModifiers
ws_sm.delete_rows(1, ws_sm.max_row)
for r in rows_to_keep:
    ws_sm.append(r)
# Add new static modifiers
ws_sm.append(['psv_ferrous_chaos', 'ATK', 'Percent', '10, 15, 20, 25, 30, 35', desc_vn])
ws_sm.append(['psv_ferrous_chaos', 'DEF', 'Percent', '10, 15, 20, 25, 30, 35', desc_vn])

# CombatEvents Sheet
ws_ce = wb_gc['CombatEvents']
ce_rows_to_keep = []
for r in list(ws_ce.iter_rows(values_only=True)):
    if r[0] != 'psv_ferrous_chaos':
        ce_rows_to_keep.append(r)
ws_ce.delete_rows(1, ws_ce.max_row)
for r in ce_rows_to_keep:
    ws_ce.append(r)

# Event 1: HP > 50% & IsSkill -> Increase Skill Damage by 16%
ws_ce.append(['psv_ferrous_chaos', 'OnBeforeDealDamage', 'Eff_IncreaseAttack', '16, 16, 16, 16, 16, 16', 'IsSkill, HP_Above_50', 0, 0, 'self', 'Trạng thái Áp Đảo (HP > 50%) tăng 16% Sát thương Kỹ năng'])
# Event 2: HP <= 50% -> 16% Lifesteal
ws_ce.append(['psv_ferrous_chaos', 'OnAfterDealDamage', 'Eff_Lifesteal', '16, 16, 16, 16, 16, 16', 'HP_BelowOrEqual_50', 0, 0, 'self', 'Trạng thái Huyết Chiến (HP ≤ 50%) nhận 16% Hút Máu'])

wb_gc.save(gc_path)
print("Updated GameConfig.xlsx with psv_ferrous_chaos")
