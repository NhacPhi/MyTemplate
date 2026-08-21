import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet EffectConfig
ws_ec = wb_gc['EffectConfig']
eff_found = False
for row in ws_ec.iter_rows(values_only=False):
    if row[0].value == 'EFF_Dragon_Speed':
        row[1].value = 'BUFF_SPEED_NAME'
        row[2].value = 'BUFF_SPEED_DES'
        row[3].value = 'StatBuff'
        row[4].value = 'SPEED'
        row[5].value = 'Percent'
        row[6].value = 2
        row[7].value = 30
        row[8].value = 1
        eff_found = True
        print("Updated EFF_Dragon_Speed in EffectConfig sheet")
        break

if not eff_found:
    ws_ec.append(['EFF_Dragon_Speed', 'BUFF_SPEED_NAME', 'BUFF_SPEED_DES', 'StatBuff', 'SPEED', 'Percent', 2, 30, 1])
    print("Added EFF_Dragon_Speed to EffectConfig sheet")

# 1.2 Sheet SkillConfig
ws_sc = wb_gc['SkillConfig']
for row in ws_sc.iter_rows(values_only=False):
    if row[0].value == 'LittleWhiteDragon_B':
        row[3].value = 'Melee'
        row[4].value = '0.8, 1, 1.2'
        row[5].value = '0, 0, 0'
        row[6].value = 'BasicAttack'
        row[7].value = 'SingleEnemy'
        row[8].value = 'None'
        print("Updated LittleWhiteDragon_B in SkillConfig")
    elif row[0].value == 'LittleWhiteDragon_M':
        row[3].value = 'ThunderBall'
        row[4].value = '1, 1.5, 1.5'
        row[5].value = '4.0, 4.0, 3.0'
        row[6].value = 'ActiveSkill'
        row[7].value = 'SingleEnemy'
        row[8].value = 'EFF_Dragon_Speed'
        print("Updated LittleWhiteDragon_M in SkillConfig")
    elif row[0].value == 'LittleWhiteDragon_U':
        row[3].value = 'EmpowerAttack'
        row[4].value = '1.5, 1.8, 1.8'
        row[5].value = '5.0, 5.0, 4.0'
        row[6].value = 'ActiveSkill'
        row[7].value = 'SingleEnemy'
        row[8].value = 'None'
        print("Updated LittleWhiteDragon_U in SkillConfig")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_battle = wb_loc['Battle']

b_vn = "Vung bảo kiếm chém dứt khoát về phía trước, gây ST bằng {0}% Tấn Công cho một kẻ địch. Đòn đánh có sẵn 30% Xuyên Giáp."
b_en = "Swings the dragon blade forward, dealing {0}% ATK to an enemy with 30% built-in Armor Penetration."

m_vn = "Triệu hồi Lôi Long ngưng tụ và bắn quả cầu sét về phía mục tiêu, gây ST bằng {0}% Tấn Công cho một kẻ địch, đồng thời bản thân lập tức nhận trạng thái [Truy Phong] - Tăng 30% Tốc Độ trong 2 hiệp."
m_en = "Summons thunder dragon energy to blast a lightning orb, dealing {0}% ATK to an enemy and grants self [Gale Chase], increasing Speed by 30% for 2 turns."

u_vn = "Vung bảo kiếm triệu hồi sóng thần cuộn trào càn quét đối phương, gây ST bằng {0}% Tấn Công cho một kẻ địch. Nếu mục tiêu đang có Giáp Ảo, sát thương gây ra được x2 (Gấp đôi)."
u_en = "Summons surging tsunami blades to sweep through the enemy, dealing {0}% ATK. Deals double damage (x2) if target currently has a Shield."

for row in ws_battle.iter_rows(values_only=False):
    if row[0].value == 'LittleWhiteDragon_B_Des':
        row[1].value = b_en
        row[2].value = b_vn
        print("Updated LittleWhiteDragon_B_Des in Localizations.xlsx")
    elif row[0].value == 'LittleWhiteDragon_M_Des':
        row[1].value = m_en
        row[2].value = m_vn
        print("Updated LittleWhiteDragon_M_Des in Localizations.xlsx")
    elif row[0].value == 'LittleWhiteDragon_U_Des':
        row[1].value = u_en
        row[2].value = u_vn
        print("Updated LittleWhiteDragon_U_Des in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
