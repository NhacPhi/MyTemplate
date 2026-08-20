import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

ws_ce = wb_gc['CombatEvents']
for row in ws_ce.iter_rows(values_only=False):
    if row[0].value == 'psv_nimbus_cudgel':
        row[1].value = 'OnAfterDealDamage'
        row[2].value = 'Eff_Action_Point_Boost'
        row[3].value = '20, 25, 30, 35, 40, 50'
        row[4].value = 'SingleTarget, TargetDead'
        row[5].value = 0
        row[7].value = 'self'
        print("Updated psv_nimbus_cudgel in CombatEvents sheet")
    elif row[0].value == 'psv_triple_edged_blade':
        row[1].value = 'OnAfterDealDamage'
        row[2].value = 'Eff_Reduce_Cooldown'
        row[3].value = '30.0, 40.0, 50.0, 60.0, 70.0, 80.0'
        row[4].value = 'IsSkill, SingleTarget, TargetDead'
        row[5].value = 1
        row[7].value = 'self'
        print("Updated psv_triple_edged_blade in CombatEvents sheet")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

found = False
for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_COOLDOWN_REDUCED':
        row[1].value = '-1 CD'
        row[2].value = '-1 Hồi Chiêu'
        found = True
        print("Updated STR_COOLDOWN_REDUCED in Localizations.xlsx")
        break

if not found:
    ws_str.append(['STR_COOLDOWN_REDUCED', '-1 CD', '-1 Hồi Chiêu'])
    print("Added STR_COOLDOWN_REDUCED to Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
