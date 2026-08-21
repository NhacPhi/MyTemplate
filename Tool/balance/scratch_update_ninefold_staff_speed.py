import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

desc_vn = 'Tăng {0}% Tốc độ. Tăng thêm {1}% hiệu quả cho các kỹ năng Cường Hóa và Buff.'
desc_en = 'Increases Speed by {0}%. Increases the effectiveness of Buff and Enhancement skills by an additional {1}%.'

# 1.1 Sheet Passives
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[1].value = 'STR_NINEFOLD_STAFF_SKILL'
        row[2].value = desc_vn
        print("Updated psv_ninefold_staff in Passives sheet")

# 1.2 Sheet StaticModifiers
ws_static = wb_gc['StaticModifiers']
for row in ws_static.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[1].value = 'SPEED'
        row[2].value = 'Percent'
        row[3].value = '6, 8, 10, 12, 15, 20'
        row[4].value = desc_vn
        print("Updated psv_ninefold_staff StaticModifiers to SPEED")

# 1.3 Sheet CombatEvents
ws_events = wb_gc['CombatEvents']
for row in ws_events.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[8].value = desc_vn
        print("Updated psv_ninefold_staff desc in CombatEvents sheet")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_NINEFOLD_STAFF_SKILL':
        row[1].value = desc_en
        row[2].value = desc_vn
        print("Updated STR_NINEFOLD_STAFF_SKILL in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
