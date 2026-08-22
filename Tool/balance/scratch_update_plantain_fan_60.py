import openpyxl
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

gc_path = 'Tool/data/GameConfig.xlsx'
loc_path = 'Tool/data/Localizations.xlsx'

desc_vi = 'Tăng {0}% Sát Thương Bạo Kích và {1}% Tấn Công. Đòn tấn công kỹ năng có 60% xác suất giảm {2}% Phòng Thủ của mục tiêu trong 2 hiệp và gây thêm {3}% Sát Thương lên kẻ địch đang chịu hiệu ứng bất lợi.'
desc_en = 'Increases CRIT DMG by {0}% and ATK by {1}%. Skill attacks have a 60% chance to reduce target DEF by {2}% for 2 turns and deal {3}% additional Damage to debuffed enemies.'

for attempt in range(5):
    try:
        wb_gc = openpyxl.load_workbook(gc_path)

        # 1. Update Passives
        ws_passives = wb_gc['Passives']
        for row in ws_passives.iter_rows(values_only=False):
            if row[0].value == 'psv_plantain_fan':
                row[1].value = 'STR_PLANTAIN_FAN_SKILL'
                row[2].value = desc_vi

        # 2. Update StaticModifiers
        ws_static = wb_gc['StaticModifiers']
        rows_to_keep = []
        for row in ws_static.iter_rows(values_only=True):
            if row[0] != 'psv_plantain_fan':
                rows_to_keep.append(row)

        ws_static.delete_rows(1, ws_static.max_row)
        for r in rows_to_keep:
            ws_static.append(r)

        ws_static.append(['psv_plantain_fan', 'CRIT_DMG', 'Percent', '20.0, 25.0, 30.0, 35.0, 40.0, 50.0'])
        ws_static.append(['psv_plantain_fan', 'ATK', 'Percent', '10.0, 12.0, 14.0, 17.0, 20.0, 25.0'])

        # 3. Update CombatEvents
        ws_events = wb_gc['CombatEvents']
        rows_events_keep = []
        for row in ws_events.iter_rows(values_only=True):
            if row[0] != 'psv_plantain_fan':
                rows_events_keep.append(row)

        ws_events.delete_rows(1, ws_events.max_row)
        for r in rows_events_keep:
            ws_events.append(r)

        ws_events.append([
            'psv_plantain_fan',
            'OnAfterDealDamage',
            'Eff_ReduceDefense',
            '15.0, 18.0, 21.0, 24.0, 27.0, 30.0',
            'IsSkill',
            60.0,
            0,
            'enemy',
            desc_vi
        ])
        ws_events.append([
            'psv_plantain_fan',
            'OnBeforeDealDamage',
            'Eff_DamageBonus',
            '15.0, 18.0, 21.0, 24.0, 27.0, 30.0',
            'TargetHasDebuff',
            0,
            0,
            'self',
            desc_vi
        ])

        wb_gc.save(gc_path)
        print("Successfully saved GameConfig.xlsx for Plantain_Fan with 60% chance!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} failed: {e}. Retrying...")
        time.sleep(1)

for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook(loc_path)
        ws_str = wb_loc['STR']
        for row in ws_str.iter_rows(values_only=False):
            if row[0].value == 'STR_PLANTAIN_FAN_SKILL':
                row[1].value = desc_en
                row[2].value = desc_vi

        wb_loc.save(loc_path)
        print("Successfully saved Localizations.xlsx for Plantain_Fan with 60% chance!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} loc failed: {e}. Retrying...")
        time.sleep(1)
