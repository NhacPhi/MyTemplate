import openpyxl
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

gc_path = 'Tool/data/GameConfig.xlsx'
loc_path = 'Tool/data/Localizations.xlsx'

desc_vi = 'Tăng {0}% Sát Thương Bạo Kích và {1}% Tấn Công. Sau khi thi triển kỹ năng tấn công đơn đánh bại mục tiêu, có {2}% xác suất hiệp hồi chiêu kỹ năng của bản thân -1.'
desc_en = 'Increases CRIT DMG by {0}% and ATK by {1}%. After defeating an enemy with a single-target skill, has a {2}% chance to reduce own skill cooldown by 1 turn.'

for attempt in range(5):
    try:
        wb_gc = openpyxl.load_workbook(gc_path)

        # 1. Update Passives
        ws_passives = wb_gc['Passives']
        for row in ws_passives.iter_rows(values_only=False):
            if row[0].value == 'psv_triple_edged_blade':
                row[1].value = 'STR_TRIPLE_EDGED_BLADE_SKILL'
                row[2].value = desc_vi

        # 2. Update StaticModifiers
        ws_static = wb_gc['StaticModifiers']
        rows_to_keep = []
        for row in ws_static.iter_rows(values_only=True):
            if row[0] != 'psv_triple_edged_blade':
                rows_to_keep.append(row)

        ws_static.delete_rows(1, ws_static.max_row)
        for r in rows_to_keep:
            ws_static.append(r)

        ws_static.append(['psv_triple_edged_blade', 'CRIT_DMG', 'Percent', '30.0, 36.0, 42.0, 48.0, 54.0, 60.0', desc_vi])
        ws_static.append(['psv_triple_edged_blade', 'ATK', 'Percent', '10.0, 12.0, 14.0, 17.0, 20.0, 25.0', desc_vi])

        # 3. Update CombatEvents
        ws_events = wb_gc['CombatEvents']
        for row in ws_events.iter_rows(values_only=False):
            if row[0].value == 'psv_triple_edged_blade':
                row[8].value = desc_vi

        wb_gc.save(gc_path)
        print("Successfully saved GameConfig.xlsx for Triple_Edged_Blade!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} failed: {e}. Retrying...")
        time.sleep(1)

for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook(loc_path)
        ws_str = wb_loc['STR']
        for row in ws_str.iter_rows(values_only=False):
            if row[0].value == 'STR_TRIPLE_EDGED_BLADE_SKILL':
                row[1].value = desc_en
                row[2].value = desc_vi

        wb_loc.save(loc_path)
        print("Successfully saved Localizations.xlsx for Triple_Edged_Blade!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} loc failed: {e}. Retrying...")
        time.sleep(1)
