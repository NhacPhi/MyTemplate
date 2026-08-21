import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Weapon
ws_weapon = wb_gc['Weapon']
for row in ws_weapon.iter_rows(values_only=False):
    if row[0].value == 'Ninefold_Staff':
        row[1].value = 'Rare'
        row[2].value = 'Support'
        row[3].value = 495
        row[4].value = 121
        row[5].value = 50
        row[6].value = 12
        row[7].value = 'psv_ninefold_staff'
        print("Updated Ninefold_Staff in Weapon sheet")

# 1.2 Sheet Passives
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[1].value = 'STR_NINEFOLD_STAFF_SKILL'
        row[2].value = 'Máu tối đa tăng {0}%. Khi thi triển kỹ năng hỗ trợ, hiệu quả tăng Tấn Công cho đồng minh được cộng thêm {1}%.'
        print("Updated psv_ninefold_staff in Passives sheet")

# 1.3 Sheet StaticModifiers
ws_static = wb_gc['StaticModifiers']
for row in ws_static.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[1].value = 'HP'
        row[2].value = 'Percent'
        row[3].value = '8, 11, 14, 17, 20, 25'
        row[4].value = 'Máu tối đa tăng {0}%. Khi thi triển kỹ năng hỗ trợ, hiệu quả tăng Tấn Công cho đồng minh được cộng thêm {1}%.'
        print("Updated psv_ninefold_staff in StaticModifiers sheet")

# 1.4 Sheet CombatEvents
ws_events = wb_gc['CombatEvents']
event_found = False
for row in ws_events.iter_rows(values_only=False):
    if row[0].value == 'psv_ninefold_staff':
        row[1].value = 'OnSkillStart'
        row[2].value = 'Eff_Buff_Enhance'
        row[3].value = '5.0, 6.0, 7.0, 8.0, 9.0, 10.0'
        row[4].value = 'None'
        row[5].value = 0
        row[6].value = 0
        row[7].value = 'self'
        row[8].value = 'Máu tối đa tăng {0}%. Khi thi triển kỹ năng hỗ trợ, hiệu quả tăng Tấn Công cho đồng minh được cộng thêm {1}%.'
        event_found = True
        print("Updated psv_ninefold_staff in CombatEvents sheet")

if not event_found:
    ws_events.append(['psv_ninefold_staff', 'OnSkillStart', 'Eff_Buff_Enhance', '5.0, 6.0, 7.0, 8.0, 9.0, 10.0', 'None', 0, 0, 'self', 'Máu tối đa tăng {0}%. Khi thi triển kỹ năng hỗ trợ, hiệu quả tăng Tấn Công cho đồng minh được cộng thêm {1}%.'])
    print("Added psv_ninefold_staff to CombatEvents sheet")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_NINEFOLD_STAFF_SKILL':
        row[1].value = 'Increases Max HP by {0}%. When casting support skills, increases ATK buff effectiveness for allies by an additional {1}%.'
        row[2].value = 'Máu tối đa tăng {0}%. Khi thi triển kỹ năng hỗ trợ, hiệu quả tăng Tấn Công cho đồng minh được cộng thêm {1}%.'
        print("Updated STR_NINEFOLD_STAFF_SKILL in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
