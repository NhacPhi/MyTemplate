import openpyxl
import sys
import time
sys.stdout.reconfigure(encoding='utf-8')

# Update GameConfig.xlsx
for attempt in range(5):
    try:
        wb_game = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')

        # 1. SkillConfig
        ws_skill = wb_game['SkillConfig']
        for row in ws_skill.iter_rows(min_row=2):
            if row[0].value == 'BullDemonKing_B':
                row[3].value = 'Melee'
                row[4].value = '1, 1.2, 1.5'
                row[5].value = '0, 0, 0'
                row[6].value = 'BasicAttack'
                row[7].value = 'SingleEnemy'
                row[8].value = 'None'
                row[9].value = 'VanguardTiger_Attack'
                row[10].value = 'psv_bulldemonking_counter'
                print('Updated SkillConfig: BullDemonKing_B')
            elif row[0].value == 'BullDemonKing_M':
                row[3].value = 'StatModifier'
                row[4].value = '0.3, 0.4, 0.5'
                row[5].value = '4.0, 4.0, 3.0'
                row[6].value = 'ActiveSkill'
                row[7].value = 'Self'
                row[8].value = 'EFF_Bull_Buff_ATK'
                row[9].value = 'BullDemonKing_Major'
                row[10].value = 'None'
                print('Updated SkillConfig: BullDemonKing_M')
            elif row[0].value == 'BullDemonKing_U':
                row[3].value = 'PoisonBall'
                row[4].value = '2.5, 3.0, 3.5'
                row[5].value = '5.0, 5.0, 4.0'
                row[6].value = 'ActiveSkill'
                row[7].value = 'SingleEnemy'
                row[8].value = 'EFF_Bull_Poison'
                row[9].value = 'BullDemonKing_Main'
                row[10].value = 'None'
                print('Updated SkillConfig: BullDemonKing_U')

        # 2. EffectConfig
        ws_eff = wb_game['EffectConfig']
        eff_keys = {row[0].value: row for row in ws_eff.iter_rows(min_row=2)}

        if 'EFF_Bull_Buff_ATK' in eff_keys:
            r = eff_keys['EFF_Bull_Buff_ATK']
            r[1].value = 'BUFF_ATTACK_NAME'
            r[2].value = 'BUFF_ATTACK_DES'
            r[3].value = 'StatBuff'
            r[4].value = 'ATK'
            r[5].value = 'Percent'
            r[6].value = 3
            r[7].value = 30
            r[8].value = 1
            print('Updated EffectConfig: EFF_Bull_Buff_ATK')
        else:
            ws_eff.append(['EFF_Bull_Buff_ATK', 'BUFF_ATTACK_NAME', 'BUFF_ATTACK_DES', 'StatBuff', 'ATK', 'Percent', 3, 30, 1])
            print('Added EffectConfig: EFF_Bull_Buff_ATK')

        if 'EFF_Bull_Poison' in eff_keys:
            r = eff_keys['EFF_Bull_Poison']
            r[1].value = 'EFF_POSION_NAME'
            r[2].value = 'EFF_POSION_DES'
            r[3].value = 'Poison'
            r[4].value = 'None'
            r[5].value = 'Percent'
            r[6].value = 3
            r[7].value = 8
            r[8].value = 1
            print('Updated EffectConfig: EFF_Bull_Poison')
        else:
            ws_eff.append(['EFF_Bull_Poison', 'EFF_POSION_NAME', 'EFF_POSION_DES', 'Poison', 'None', 'Percent', 3, 8, 1])
            print('Added EffectConfig: EFF_Bull_Poison')

        # 3. Passives
        ws_pass = wb_game['Passives']
        pass_keys = {row[0].value: row for row in ws_pass.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in pass_keys:
            ws_pass.append(['psv_bulldemonking_counter', 'STR_BULL_COUNTER_NAME', 'Có 50%, 75%, 100% cơ hội phản kích khi bị đánh thường', 'STR_BULL_COUNTER_DES'])
            print('Added psv_bulldemonking_counter to Passives')
        else:
            r = pass_keys['psv_bulldemonking_counter']
            r[1].value = 'STR_BULL_COUNTER_NAME'
            r[2].value = 'Có 50%, 75%, 100% cơ hội phản kích khi bị đánh thường'
            r[3].value = 'STR_BULL_COUNTER_DES'
            print('Updated psv_bulldemonking_counter in Passives')

        # 4. CombatEvents
        ws_ce = wb_game['CombatEvents']
        ce_keys = {row[0].value: row for row in ws_ce.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in ce_keys:
            ws_ce.append(['psv_bulldemonking_counter', 'OnTakeDamage', 'Eff_CounterAttack', 'BasicAttack', 'enemy', '50, 75, 100, 100, 100, 100'])
            print('Added psv_bulldemonking_counter to CombatEvents')
        else:
            r = ce_keys['psv_bulldemonking_counter']
            r[1].value = 'OnTakeDamage'
            r[2].value = 'Eff_CounterAttack'
            r[3].value = 'BasicAttack'
            r[4].value = 'enemy'
            r[5].value = '50, 75, 100, 100, 100, 100'
            print('Updated psv_bulldemonking_counter in CombatEvents')

        wb_game.save('Tool/data/GameConfig.xlsx')
        print('Successfully saved GameConfig.xlsx!')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: GameConfig.xlsx is locked, retrying in 1s...')
        time.sleep(1)
