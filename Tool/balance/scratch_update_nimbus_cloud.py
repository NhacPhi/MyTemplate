import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# Check and update Passives sheet
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_nimbus_cloud':
        row[1].value = 'STR_NIMBUS_CLOUD_SKILL'
        row[2].value = 'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'
        print("Updated Passives sheet for psv_nimbus_cloud")

# Update StaticModifiers sheet
ws_static = wb_gc['StaticModifiers']
# Remove old rows for psv_nimbus_cloud if any
rows_to_keep = []
for row in ws_static.iter_rows(values_only=True):
    if row[0] != 'psv_nimbus_cloud':
        rows_to_keep.append(row)

# Clear and rewrite with new modifiers
ws_static.delete_rows(1, ws_static.max_row)
for r in rows_to_keep:
    ws_static.append(r)

# Append new static modifiers
ws_static.append(['psv_nimbus_cloud', 'SPEED', 'Percent', '8.0, 10.0, 12.0, 14.0, 16.0, 20.0'])
ws_static.append(['psv_nimbus_cloud', 'ATK', 'Percent', '10.0, 12.0, 14.0, 17.0, 20.0, 25.0'])
print("Added StaticModifiers for psv_nimbus_cloud")

# Update CombatEvents sheet
ws_events = wb_gc['CombatEvents']
rows_events_keep = []
for row in ws_events.iter_rows(values_only=True):
    if row[0] != 'psv_nimbus_cloud':
        rows_events_keep.append(row)

ws_events.delete_rows(1, ws_events.max_row)
for r in rows_events_keep:
    ws_events.append(r)

ws_events.append([
    'psv_nimbus_cloud',
    'OnTurnStart',
    'Eff_Action_Point_Boost',
    '15.0, 18.0, 21.0, 24.0, 28.0, 35.0',
    None,
    0,
    0,
    'self',
    'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'
])
print("Updated CombatEvents for psv_nimbus_cloud")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_NIMBUS_CLOUD_SKILL':
        row[1].value = 'Increases SPEED by {0}% and ATK by {1}%. When turn starts, increases self Action Point by {2}%.'
        row[2].value = 'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'
        print("Updated STR_NIMBUS_CLOUD_SKILL in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
