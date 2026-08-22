import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

desc_vn = 'Tăng {0}% Sát Thương Bạo Kích và {1}% Tấn Công. Khi gây sát thương Bạo Kích, tăng {2}% Điểm Hành Động của bản thân. Sau khi tấn công đơn mục tiêu, nếu mục tiêu còn sống, có {3}% xác suất lập tức tung đòn Đánh Thường để truy kích.'
desc_en = 'Increases CRIT DMG by {0}% and ATK by {1}%. Landing a Critical Hit increases self Action Point by {2}%. After a single-target attack, if the target is still alive, has a {3}% chance to perform a Basic Attack as follow-up.'

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Passives
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_vanguard_of_volition':
        row[1].value = 'STR_VANGUARD_OF_VOLITION_SKILL'
        row[2].value = desc_vn
        print("Updated psv_vanguard_of_volition in Passives sheet")

# 1.2 Sheet StaticModifiers
ws_static = wb_gc['StaticModifiers']
# Remove existing psv_vanguard_of_volition rows
rows_to_del = []
for idx, row in enumerate(ws_static.iter_rows(values_only=True), start=1):
    if row[0] == 'psv_vanguard_of_volition':
        rows_to_del.append(idx)

for r_idx in reversed(rows_to_del):
    ws_static.delete_rows(r_idx)

# Append new static modifiers
ws_static.append(['psv_vanguard_of_volition', 'CRIT_DMG', 'Percent', '20, 25, 30, 35, 40, 50', desc_vn])
ws_static.append(['psv_vanguard_of_volition', 'ATK', 'Percent', '10, 12, 14, 17, 20, 25', desc_vn])
print("Added StaticModifiers for psv_vanguard_of_volition (CRIT_DMG & ATK)")

# 1.3 Sheet CombatEvents
ws_events = wb_gc['CombatEvents']
events_to_del = []
for idx, row in enumerate(ws_events.iter_rows(values_only=True), start=1):
    if row[0] == 'psv_vanguard_of_volition':
        events_to_del.append(idx)

for r_idx in reversed(events_to_del):
    ws_events.delete_rows(r_idx)

# Append new combat events
# Event 1: Action Point Boost on Crit
ws_events.append(['psv_vanguard_of_volition', 'OnAfterDealDamage', 'Eff_Action_Point_Boost', '15.0, 18.0, 21.0, 24.0, 27.0, 30.0', 'IsCritical', 0, 0, 'self', desc_vn])
# Event 2: Follow-up basic attack on single target
ws_events.append(['psv_vanguard_of_volition', 'OnAfterSkillExcute', 'Eff_FollowUp_BasicAttack', '50.0, 60.0, 70.0, 80.0, 90.0, 100.0', 'SingleTarget, TargetAlive', 0, 0, 'self', desc_vn])
print("Added CombatEvents for psv_vanguard_of_volition (Action Point Boost & Follow-up Attack)")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)

# 2.1 Sheet STR
ws_str = wb_loc['STR']
for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_VANGUARD_OF_VOLITION_NAME':
        row[1].value = 'Vanguard of Volition'
        row[2].value = 'Tùy Tâm Thiết Can Binh'
    elif row[0].value == 'STR_VANGUARD_OF_VOLITION_DES':
        row[1].value = "The miraculous staff of the Six-Eared Macaque weighing 13,500 kilograms.\\n Though not forged in Laozi's Eight Trigram Furnace, it possesses mighty power comparable to the Nimbus Cudgel.\\n Both are divine weapons capable of infinite transformations, indistinguishable from the real."
        row[2].value = "Binh khí thần kỳ của Lục Nhĩ Mỹ Hầu nặng một vạn ba ngàn năm trăm cân.\\n Tuy không được luyện từ lò Bát Quái của Thái Thượng Lão Quân, nhưng cây gậy này có uy lực kinh thiên không hề thua kém Gậy Như Ý.\\n Cả hai món đều biến hóa khôn lường, thật giả khó phân."
    elif row[0].value == 'STR_VANGUARD_OF_VOLITION_SKILL':
        row[1].value = desc_en
        row[2].value = desc_vn
        print("Updated STR_VANGUARD_OF_VOLITION_SKILL")

# 2.2 Sheet Battle (Add STR_PURSUIT_ATTACK if not present)
ws_battle = wb_loc['Battle']
pursuit_found = False
for row in ws_battle.iter_rows(values_only=False):
    if row[0].value == 'STR_PURSUIT_ATTACK':
        row[1].value = 'Pursuit!'
        row[2].value = 'Truy Kích!'
        pursuit_found = True
        break

if not pursuit_found:
    ws_battle.append(['STR_PURSUIT_ATTACK', 'Pursuit!', 'Truy Kích!'])
    print("Added STR_PURSUIT_ATTACK to Battle sheet")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
