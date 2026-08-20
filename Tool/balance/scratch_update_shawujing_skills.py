import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet EffectConfig
ws_ec = wb_gc['EffectConfig']
eff_found = False
for row in ws_ec.iter_rows(values_only=False):
    if row[0].value == 'EFF_Sha_Buff_DEF':
        row[1].value = 'BUFF_DEF_NAME'
        row[2].value = 'BUFF_DEF_DES'
        row[3].value = 'StatBuff'
        row[4].value = 'DEF'
        row[5].value = 'Percent'
        row[6].value = 2
        row[7].value = 25
        row[8].value = 1
        eff_found = True
        print("Updated EFF_Sha_Buff_DEF in EffectConfig sheet")
        break

if not eff_found:
    ws_ec.append(['EFF_Sha_Buff_DEF', 'BUFF_DEF_NAME', 'BUFF_DEF_DES', 'StatBuff', 'DEF', 'Percent', 2, 25, 1])
    print("Added EFF_Sha_Buff_DEF to EffectConfig sheet")

# 1.2 Sheet SkillConfig
ws_sc = wb_gc['SkillConfig']
for row in ws_sc.iter_rows(values_only=False):
    if row[0].value == 'ShaWujing_B':
        row[4].value = '0.8, 1, 1.2'
        row[5].value = '0, 0, 0'
        row[6].value = 'BasicAttack'
        row[7].value = 'SingleEnemy'
        row[8].value = 'None'
        print("Updated ShaWujing_B in SkillConfig")
    elif row[0].value == 'ShaWujing_M':
        row[4].value = '1.0, 1.3, 1.6'
        row[5].value = '4.0, 4.0, 3.0'
        row[6].value = 'NonAttackSkill'
        row[7].value = 'Self'
        row[8].value = 'EFF_Sha_Buff_DEF'
        print("Updated ShaWujing_M in SkillConfig")
    elif row[0].value == 'ShaWujing_U':
        row[4].value = '1.4, 1.6, 1.8'
        row[5].value = '5.0, 5.0, 4.0'
        row[6].value = 'ActiveSkill'
        row[7].value = 'SingleEnemy'
        row[8].value = 'EFF_DefSpd_def'
        print("Updated ShaWujing_U in SkillConfig")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_battle = wb_loc['Battle']

m_vn = "Tu một ngụm linh dược từ hồ lô, lập tức hồi phục lượng Máu tương đương {0}% Tấn Công cho bản thân và tăng 25% Phòng Thủ trong 2 hiệp."
m_en = "Takes a sip of divine elixir from the gourd, immediately restoring HP equal to {0}% ATK and increasing DEF by 25% for 2 turns."

u_vn = "Xoay mạnh Huyền Trượng rồi giáng một đòn cực uy lực xuống mục tiêu, gây ST bằng {0}% Tấn Công cho một kẻ địch, đồng thời khiến kẻ địch rơi vào trạng thái [Cát Lún], giảm 20% Tốc Độ trong 2 hiệp."
u_en = "Swings the Divine Staff and slams down with extreme force, dealing {0}% ATK damage to an enemy and inflicting [Quicksand], reducing target's Speed by 20% for 2 turns."

for row in ws_battle.iter_rows(values_only=False):
    if row[0].value == 'ShaWujing_M_Des':
        row[1].value = m_en
        row[2].value = m_vn
        print("Updated ShaWujing_M_Des in Localizations.xlsx")
    elif row[0].value == 'ShaWujing_U_Des':
        row[1].value = u_en
        row[2].value = u_vn
        print("Updated ShaWujing_U_Des in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
