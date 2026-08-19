import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Weapon
ws_w = wb_gc['Weapon']
w_found = False
for row in ws_w.iter_rows(values_only=False):
    if row[0].value == 'Windfire_Wheel':
        row[1].value = 'Epic'
        row[2].value = 'Assassin'
        row[3].value = 578
        row[4].value = 214
        row[5].value = 58
        row[6].value = 21
        row[7].value = 'psv_windfire_wheel'
        w_found = True
        print("Updated Windfire_Wheel in Weapon sheet to Epic Assassin stats")
        break
if not w_found:
    ws_w.append(['Windfire_Wheel', 'Epic', 'Assassin', 578, 214, 58, 21, 'psv_windfire_wheel'])

# 1.2 Sheet Passives
ws_p = wb_gc['Passives']
desc_str = "Tăng {0}% Sát Thương Bạo Kích và {0}% Tấn Công. Khi gây sát thương Bạo Kích, tăng {1}% Điểm Hành Động của bản thân."
p_found = False
for row in ws_p.iter_rows(values_only=False):
    if row[0].value == 'psv_windfire_wheel':
        row[1].value = 'STR_WINDFIRE_WHEEL_SKILL'
        row[2].value = desc_str
        p_found = True
        print("Updated psv_windfire_wheel in Passives sheet")
        break
if not p_found:
    ws_p.append(['psv_windfire_wheel', 'STR_WINDFIRE_WHEEL_SKILL', desc_str])

# 1.3 Sheet StaticModifiers
ws_sm = wb_gc['StaticModifiers']
sm_rows = []
for r in list(ws_sm.iter_rows(values_only=True)):
    if r[0] != 'psv_windfire_wheel':
        sm_rows.append(r)
ws_sm.delete_rows(1, ws_sm.max_row)
for r in sm_rows:
    ws_sm.append(r)

ws_sm.append(['psv_windfire_wheel', 'CRIT_DMG', 'Percent', '8, 10, 12, 14, 17, 20', desc_str])
ws_sm.append(['psv_windfire_wheel', 'ATK', 'Percent', '8, 10, 12, 14, 17, 20', desc_str])
print("Updated StaticModifiers for psv_windfire_wheel (CRIT_DMG & ATK)")

# 1.4 Sheet CombatEvents
ws_ce = wb_gc['CombatEvents']
ce_rows = []
for r in list(ws_ce.iter_rows(values_only=True)):
    if r[0] != 'psv_windfire_wheel':
        ce_rows.append(r)
ws_ce.delete_rows(1, ws_ce.max_row)
for r in ce_rows:
    ws_ce.append(r)

ws_ce.append(['psv_windfire_wheel', 'OnAfterDealDamage', 'Eff_Action_Point_Boost', '10, 12, 14, 16, 18, 20', 'IsCritical', 0, 0, 'self', 'Khi gây sát thương Bạo Kích, tăng {1}% Điểm Hành Động của bản thân.'])
print("Updated CombatEvents for psv_windfire_wheel")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_loc = wb_loc['STR']

vn_desc = "Tăng {0}% Sát Thương Bạo Kích và {0}% Tấn Công. Khi gây sát thương Bạo Kích, tăng {1}% Điểm Hành Động của bản thân."
en_desc = "Increases Crit DMG and ATK by {0}%. Landing a Critical Hit increases self Action Point by {1}%. "

found = False
for row in ws_loc.iter_rows(values_only=False):
    if row[0].value == 'STR_WINDFIRE_WHEEL_SKILL':
        row[1].value = en_desc
        row[2].value = vn_desc
        found = True
        break
if not found:
    ws_loc.append(['STR_WINDFIRE_WHEEL_SKILL', en_desc, vn_desc])

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
