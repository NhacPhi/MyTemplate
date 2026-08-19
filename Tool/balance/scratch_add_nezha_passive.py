import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# SkillConfig Sheet
ws_sc = wb_gc['SkillConfig']
for row in ws_sc.iter_rows(values_only=False):
    if row[0].value == 'ThirdPrinceNezha_U':
        row[10].value = 'psv_nezha_ultimate'
        print("Updated ThirdPrinceNezha_U Passive_ID to psv_nezha_ultimate")
        break

# Passives Sheet
ws_p = wb_gc['Passives']
p_found = False
for row in ws_p.iter_rows(values_only=False):
    if row[0].value == 'psv_nezha_ultimate':
        row[1].value = 'STR_NEZHA_ULT_PASSIVE'
        row[2].value = 'Khi kết liễu kẻ địch bằng chiêu cuối, kích hoạt hiệu ứng [Càn Quét] tấn công 1 kẻ địch khác với 120% ATK.'
        p_found = True
        break
if not p_found:
    ws_p.append(['psv_nezha_ultimate', 'STR_NEZHA_ULT_PASSIVE', 'Khi kết liễu kẻ địch bằng chiêu cuối, kích hoạt hiệu ứng [Càn Quét] tấn công 1 kẻ địch khác với 120% ATK.'])
    print("Added psv_nezha_ultimate to Passives sheet")

# CombatEvents Sheet
ws_ce = wb_gc['CombatEvents']
ce_rows = []
for r in list(ws_ce.iter_rows(values_only=True)):
    if r[0] != 'psv_nezha_ultimate':
        ce_rows.append(r)
ws_ce.delete_rows(1, ws_ce.max_row)
for r in ce_rows:
    ws_ce.append(r)

ws_ce.append(['psv_nezha_ultimate', 'OnAfterDealDamage', 'Eff_FollowUp_BasicAttack', '120, 120, 120, 120, 120, 120', 'UltimateSkill, TargetDead', 0, 0, 'self', 'Khi kết liễu kẻ địch bằng chiêu cuối, kích hoạt hiệu ứng [Càn Quét] tấn công 1 kẻ địch khác với 120% ATK.'])
print("Added psv_nezha_ultimate to CombatEvents sheet")

wb_gc.save(gc_path)

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_loc = wb_loc['Battle']

vn_text = "Khi kết liễu kẻ địch bằng chiêu cuối, kích hoạt hiệu ứng [Càn Quét] tấn công 1 kẻ địch khác với 120% ATK."
en_text = "When defeating an enemy with Ultimate, activates [Battlefield Sweep], attacking another enemy for 120% ATK."

found = False
for row in ws_loc.iter_rows(values_only=False):
    if row[0].value == 'STR_NEZHA_ULT_PASSIVE':
        row[1].value = en_text
        row[2].value = vn_text
        found = True
        break

if not found:
    ws_loc.append(['STR_NEZHA_ULT_PASSIVE', en_text, vn_text])
    print("Added STR_NEZHA_ULT_PASSIVE to Localizations.xlsx")

wb_loc.save(loc_path)
