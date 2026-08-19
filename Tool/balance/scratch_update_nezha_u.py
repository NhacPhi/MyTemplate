import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)
ws_sc = wb_gc['SkillConfig']

for row in ws_sc.iter_rows(values_only=False):
    if row[0].value == 'ThirdPrinceNezha_U':
        row[3].value = 'NezhaExecution'
        print("Updated ThirdPrinceNezha_U to skill: NezhaExecution")
        break

wb_gc.save(gc_path)

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_loc = wb_loc['Battle']

vn_desc = "Áp sát mục tiêu, thúc động Hỏa Tiêm Thương bộc phát phun ra luồng lửa bão táp thiêu rụi đối phương, gây ST bằng {0}% Tấn Công cho một kẻ địch. Nếu đòn đánh KẾT LIỄU mục tiêu, kích hoạt hiệu ứng [Càn Quét], tự động truy kích 1 kẻ địch khác với 120% Tấn Công."
en_desc = "Strikes an enemy dealing {0}% ATK. If this skill KILLS the target, activates [Battlefield Sweep], automatically chasing and attacking another enemy for 120% ATK."

found = False
for row in ws_loc.iter_rows(values_only=False):
    if row[0].value == 'ThirdPrinceNezha_U_Des':
        row[1].value = en_desc
        row[2].value = vn_desc
        found = True
        break

if not found:
    ws_loc.append(['ThirdPrinceNezha_U_Des', en_desc, vn_desc])

wb_loc.save(loc_path)
print("Updated Localizations.xlsx for ThirdPrinceNezha_U_Des")
