import openpyxl, sys, re

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')
wb_loc = openpyxl.load_workbook('Tool/data/Localizations.xlsx')

loc_dict = {}
for r in wb_loc['STR'].iter_rows(values_only=True):
    if r[0]:
        loc_dict[r[0]] = (r[1], r[2])

passives = {}
for r in wb['Passives'].iter_rows(values_only=True):
    if r[0] and r[0] != 'ID':
        passives[r[0]] = {'key': r[1], 'vn': r[2], 'static': [], 'events': []}

for r in wb['StaticModifiers'].iter_rows(values_only=True):
    if r[0] in passives:
        passives[r[0]]['static'].append((r[1], r[3]))

for r in wb['CombatEvents'].iter_rows(values_only=True):
    if r[0] in passives:
        passives[r[0]]['events'].append((r[1], r[2], r[3]))

for p_id, p in passives.items():
    vn_text = loc_dict.get(p['key'], (None, p['vn']))[1]
    placeholders = re.findall(r'\{(\d+)\}', str(vn_text))
    k = p['key']
    print(f"=== {p_id} ({k}) ===")
    print(f"  Text: {vn_text}")
    print(f"  Placeholders in text: {placeholders}")
    print(f"  Static count: {len(p['static'])}, Events count: {len(p['events'])}")
    for s in p['static']:
        print(f"    Static: {s}")
    for e in p['events']:
        print(f"    Event: {e}")
