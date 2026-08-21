import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Passives
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_sunburst_spade':
        row[1].value = 'STR_SUNBURST_SPADE_SKILL'
        row[2].value = 'Tấn công tăng {0}%.'
    elif row[0].value == 'psv_bronze_hammer':
        row[1].value = 'STR_BRONZE_HAMMER_SKILL'
        row[2].value = 'Phòng thủ tăng {0}%.'
    elif row[0].value == 'psv_ennead_spear':
        row[1].value = 'STR_ENNEAD_SPEAR_SKILL'
        row[2].value = 'Tỷ lệ bạo kích tăng {0}%.'

# 1.2 Sheet StaticModifiers
ws_static = wb_gc['StaticModifiers']
sunburst_found = False
for row in ws_static.iter_rows(values_only=False):
    if row[0].value == 'psv_sunburst_spade':
        row[1].value = 'ATK'
        row[2].value = 'Percent'
        row[3].value = '5, 7.5, 10, 12.5, 15, 20'
        row[4].value = 'Tấn công tăng {0}%.'
        sunburst_found = True
    elif row[0].value == 'psv_bronze_hammer':
        row[1].value = 'DEF'
        row[2].value = 'Percent'
        row[3].value = '5, 7.5, 10, 12.5, 16, 20'
        row[4].value = 'Phòng thủ tăng {0}%.'
    elif row[0].value == 'psv_ennead_spear':
        row[1].value = 'CRIT_RATE'
        row[2].value = 'Percent'
        row[3].value = '5, 7.5, 10, 12.5, 16, 20'
        row[4].value = 'Tỷ lệ bạo kích tăng {0}%.'

if not sunburst_found:
    ws_static.append(['psv_sunburst_spade', 'ATK', 'Percent', '5, 7.5, 10, 12.5, 15, 20', 'Tấn công tăng {0}%.'])
    print("Added psv_sunburst_spade to StaticModifiers")

# 1.3 Sheet CombatEvents (Remove psv_sunburst_spade from CombatEvents)
ws_events = wb_gc['CombatEvents']
rows_to_delete = []
for idx, row in enumerate(ws_events.iter_rows(values_only=True), start=1):
    if row[0] == 'psv_sunburst_spade':
        rows_to_delete.append(idx)

for r_idx in reversed(rows_to_delete):
    ws_events.delete_rows(r_idx)
    print(f"Deleted psv_sunburst_spade from CombatEvents row {r_idx}")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

loc_updates = {
    'STR_SUNBURST_SPADE_NAME': ('Sunburst Spade', 'Xẻng Tứ Minh'),
    'STR_SUNBURST_SPADE_DES': (
        'The weapon of Yellow Lion who serves under Ninefold Primosaint.\\n It can smash giant boulders and excavate mountains.',
        'Binh khí của Hoàng Sư Tinh dưới trướng Cửu Linh Nguyên Thánh. Có thể đập vỡ đá tảng và đào hang xẻ núi.'
    ),
    'STR_SUNBURST_SPADE_SKILL': (
        'Increases ATK by {0}%.',
        'Tấn công tăng {0}%.'
    ),
    'STR_BRONZE_HAMMER_NAME': ('Bronze Hammer', 'Trùy Đồng'),
    'STR_BRONZE_HAMMER_DES': (
        'The weapon of Baize Demon who serves under Ninefold Primosaint.\\n Although it named itself after the divine beast Baize, it is actually just an ordinary monster.',
        'Binh khí của Bạch Trạch Tinh dưới trướng Cửu Linh Nguyên Thánh. Tuy tự xưng là Bạch Trạch nhưng thực chất chỉ là một yêu linh bình thường.'
    ),
    'STR_BRONZE_HAMMER_SKILL': (
        'Increases DEF by {0}%.',
        'Phòng thủ tăng {0}%.'
    ),
    'STR_ENNEAD_SPEAR_NAME': ('Ennead Spear', 'Xiên Chín Đầu'),
    'STR_ENNEAD_SPEAR_DES': (
        'This mighty divine weapon kept in the Dragon Palace of the Eastern Sea weighs 1,800 kilograms.\\n The Monkey King once remarked: "Light, too light!" and "Not handy at all!"',
        'Binh khí thần được Đông Hải Long Cung thu thập, nặng ba ngàn sáu trăm cân.\\n Mỹ Hầu Vương từng bình phẩm: "Nhẹ ơi là nhẹ!" rồi lại nói: "Chẳng tiện tay tí nào!"'
    ),
    'STR_ENNEAD_SPEAR_SKILL': (
        'Increases CRIT Rate by {0}%.',
        'Tỷ lệ bạo kích tăng {0}%.'
    )
}

for row in ws_str.iter_rows(values_only=False):
    key = row[0].value
    if key in loc_updates:
        row[1].value = loc_updates[key][0]
        row[2].value = loc_updates[key][1]
        print(f"Updated localization {key}")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
