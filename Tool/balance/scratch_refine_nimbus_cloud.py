import openpyxl
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

gc_path = 'Tool/data/GameConfig.xlsx'
loc_path = 'Tool/data/Localizations.xlsx'

for attempt in range(5):
    try:
        wb_gc = openpyxl.load_workbook(gc_path)

        # Update Passives sheet
        ws_passives = wb_gc['Passives']
        for row in ws_passives.iter_rows(values_only=False):
            if row[0].value == 'psv_nimbus_cloud':
                row[1].value = 'STR_NIMBUS_CLOUD_SKILL'
                row[2].value = 'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'

        # Update StaticModifiers sheet
        ws_static = wb_gc['StaticModifiers']
        rows_to_keep = []
        for row in ws_static.iter_rows(values_only=True):
            if row[0] != 'psv_nimbus_cloud':
                rows_to_keep.append(row)

        ws_static.delete_rows(1, ws_static.max_row)
        for r in rows_to_keep:
            ws_static.append(r)

        ws_static.append(['psv_nimbus_cloud', 'SPEED', 'Percent', '4.0, 5.0, 6.0, 7.0, 8.5, 10.0'])
        ws_static.append(['psv_nimbus_cloud', 'ATK', 'Percent', '5.0, 6.5, 8.0, 10.0, 12.0, 15.0'])

        # Update CombatEvents sheet
        ws_events = wb_gc['CombatEvents']
        rows_events_keep = []
        for row in ws_events.iter_rows(values_only=True):
            if row[0] != 'psv_nimbus_cloud':
                rows_events_keep.append(row)

        ws_events.delete_rows(1, ws_events.max_row)
        for r in rows_events_keep:
            ws_events.append(r)

        ws_events.append([
            'psv_nimbus_cloud',
            'OnTurnStart',
            'Eff_Action_Point_Boost',
            '10.0, 12.0, 14.0, 16.0, 18.0, 20.0',
            None,
            0,
            0,
            'self',
            'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'
        ])

        wb_gc.save(gc_path)
        print("Successfully saved GameConfig.xlsx!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} failed: {e}. Retrying in 1s...")
        time.sleep(1)

wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']
for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_NIMBUS_CLOUD_SKILL':
        row[1].value = 'Increases SPEED by {0}% and ATK by {1}%. When turn starts, increases self Action Point by {2}%.'
        row[2].value = 'Tăng {0}% Tốc Độ và {1}% Tấn Công. Khi bắt đầu lượt, tăng {2}% Điểm Hành Động của bản thân.'

wb_loc.save(loc_path)
print("Successfully saved Localizations.xlsx!")
