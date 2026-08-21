import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet EffectConfig
ws_ec = wb_gc['EffectConfig']
eff_found = False
for row in ws_ec.iter_rows(values_only=False):
    if row[0].value == 'EFF_Tang_Buff_ATK':
        row[1].value = 'BUFF_ATTACK_NAME'
        row[2].value = 'BUFF_ATTACK_DES'
        row[3].value = 'StatBuff'
        row[4].value = 'ATK'
        row[5].value = 'Percent'
        row[6].value = 2
        row[7].value = 20
        row[8].value = 1
        eff_found = True
        print("Updated EFF_Tang_Buff_ATK in EffectConfig sheet")
        break

if not eff_found:
    ws_ec.append(['EFF_Tang_Buff_ATK', 'BUFF_ATTACK_NAME', 'BUFF_ATTACK_DES', 'StatBuff', 'ATK', 'Percent', 2, 20, 1])
    print("Added EFF_Tang_Buff_ATK to EffectConfig sheet")

# 1.2 Sheet SkillConfig
ws_sc = wb_gc['SkillConfig']
for row in ws_sc.iter_rows(values_only=False):
    if row[0].value == 'TangSanZang_B':
        row[3].value = 'Range'
        row[4].value = '0.8, 1, 1.2'
        row[5].value = '0, 0, 0'
        row[6].value = 'BasicAttack'
        row[7].value = 'SingleEnemy'
        row[8].value = 'None'
        print("Updated TangSanZang_B in SkillConfig")
    elif row[0].value == 'TangSanZang_M':
        row[3].value = 'StatModifier'
        row[4].value = '0, 0, 0'
        row[5].value = '4.0, 4.0, 3.0'
        row[6].value = 'NonAttackSkill'
        row[7].value = 'SameRowAllies'
        row[8].value = 'EFF_Tang_Buff_ATK'
        print("Updated TangSanZang_M in SkillConfig")
    elif row[0].value == 'TangSanZang_U':
        row[3].value = 'EmpowerAttack'
        row[4].value = '1.5, 1.8, 1.8'
        row[5].value = '5.0, 5.0, 4.0'
        row[6].value = 'ActiveSkill'
        row[7].value = 'SingleEnemy'
        row[8].value = 'EFF_Stun_def'
        print("Updated TangSanZang_U in SkillConfig")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_battle = wb_loc['Battle']

m_vn = "Ngưng tụ màn kinh thư vàng bao bọc đồng đội, tăng 20% Tấn Công cho toàn bộ đồng minh cùng hàng trong 2 hiệp."
m_en = "Condenses golden scripture barrier around allies, increasing ATK by 20% for all allies in the same row for 2 turns."

u_vn = "Áp sát kẻ địch, đưa Bát Vàng ra phóng chiếu luồng hào quang Phật pháp chói lọi, gây ST bằng {0}% Tấn Công cho một kẻ địch, đồng thời khiến mục tiêu bị [Choáng] trong 1 hiệp."
u_en = "Approaches the enemy and shines dazzling Buddhist radiance from the Golden Bowl, dealing {0}% ATK damage to an enemy and inflicting [Stun] for 1 turn."

for row in ws_battle.iter_rows(values_only=False):
    if row[0].value == 'TangSanZang_M_Des':
        row[1].value = m_en
        row[2].value = m_vn
        print("Updated TangSanZang_M_Des in Localizations.xlsx")
    elif row[0].value == 'TangSanZang_U_Des':
        row[1].value = u_en
        row[2].value = u_vn
        print("Updated TangSanZang_U_Des in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
