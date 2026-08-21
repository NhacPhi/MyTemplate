import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Weapon
ws_weapon = wb_gc['Weapon']
for row in ws_weapon.iter_rows(values_only=False):
    if row[0].value == 'Reincarnation':
        row[1].value = 'Epic'
        row[2].value = 'ADCarry'
        row[3].value = 578
        row[4].value = 214
        row[5].value = 58
        row[6].value = 21
        row[7].value = 'psv_reincarnation'
        print("Updated Reincarnation in Weapon sheet")

# 1.2 Sheet Passives
ws_passives = wb_gc['Passives']
for row in ws_passives.iter_rows(values_only=False):
    if row[0].value == 'psv_reincarnation':
        row[1].value = 'STR_REINCARNATION_SKILL'
        row[2].value = 'Tấn công tăng {0}%. Đòn tấn công đơn mục tiêu gây thêm {1}% Sát thương.'
        print("Updated psv_reincarnation in Passives sheet")

# 1.3 Sheet StaticModifiers
ws_static = wb_gc['StaticModifiers']
for row in ws_static.iter_rows(values_only=False):
    if row[0].value == 'psv_reincarnation':
        row[1].value = 'ATK'
        row[2].value = 'Percent'
        row[3].value = '8, 11, 14, 17, 20, 25'
        row[4].value = 'Tấn công tăng {0}%. Đòn tấn công đơn mục tiêu gây thêm {1}% Sát thương.'
        print("Updated psv_reincarnation in StaticModifiers sheet")

# 1.4 Sheet CombatEvents
ws_events = wb_gc['CombatEvents']
event_found = False
for row in ws_events.iter_rows(values_only=False):
    if row[0].value == 'psv_reincarnation':
        row[1].value = 'OnBeforeDealDamage'
        row[2].value = 'Eff_IncreaseAttack'
        row[3].value = '10.0, 13.0, 16.0, 19.0, 22.0, 25.0'
        row[4].value = 'SingleTarget'
        row[5].value = 0
        row[6].value = 0
        row[7].value = 'self'
        row[8].value = 'Tấn công tăng {0}%. Đòn tấn công đơn mục tiêu gây thêm {1}% Sát thương.'
        event_found = True
        print("Updated psv_reincarnation in CombatEvents sheet")

if not event_found:
    ws_events.append(['psv_reincarnation', 'OnBeforeDealDamage', 'Eff_IncreaseAttack', '10.0, 13.0, 16.0, 19.0, 22.0, 25.0', 'SingleTarget', 0, 0, 'self', 'Tấn công tăng {0}%. Đòn tấn công đơn mục tiêu gây thêm {1}% Sát thương.'])
    print("Added psv_reincarnation to CombatEvents sheet")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_REINCARNATION_SKILL':
        row[1].value = 'Increases ATK by {0}%. Single-target attacks deal an additional {1}% Damage.'
        row[2].value = 'Tấn công tăng {0}%. Đòn tấn công đơn mục tiêu gây thêm {1}% Sát thương.'
        print("Updated STR_REINCARNATION_SKILL in Localizations.xlsx")

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
