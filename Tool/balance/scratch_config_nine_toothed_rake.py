import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

desc_str_vn = "Tăng {0}% HP và {0}% Phòng Thủ. Mỗi khi chịu sát thương, Phòng Thủ tăng thêm {1}%, tối đa cộng dồn 5 tầng. Khi đạt 5 tầng, hồi phục 10% HP Tối Đa."
desc_str_en = "Increases HP and DEF by {0}%. Whenever taking damage, DEF increases by {1}%, stacking up to 5 times. Upon reaching 5 stacks, restores 10% Max HP."

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet Weapon
ws_w = wb_gc['Weapon']
for row in ws_w.iter_rows(values_only=False):
    if row[0].value == 'Nine_ToolThed_Rake':
        row[1].value = 'Legendary'
        row[2].value = 'Tanker'
        row[3].value = 2000
        row[4].value = 100
        row[5].value = 200
        row[6].value = 10
        row[7].value = 'psv_nine_toothed_rake'
        print("Updated Nine_ToolThed_Rake in Weapon sheet")
        break

# 1.2 Sheet Passives
ws_p = wb_gc['Passives']
p_found = False
for row in ws_p.iter_rows(values_only=False):
    if row[0].value == 'psv_nine_toothed_rake':
        row[1].value = 'STR_NINE_TOOLTHED_RAKE_SKILL'
        row[2].value = desc_str_vn
        p_found = True
        print("Updated psv_nine_toothed_rake in Passives sheet")
        break
if not p_found:
    ws_p.append(['psv_nine_toothed_rake', 'STR_NINE_TOOLTHED_RAKE_SKILL', desc_str_vn])

# 1.3 Sheet StaticModifiers
ws_sm = wb_gc['StaticModifiers']
sm_rows = []
for r in list(ws_sm.iter_rows(values_only=True)):
    if r[0] != 'psv_nine_toothed_rake':
        sm_rows.append(r)
ws_sm.delete_rows(1, ws_sm.max_row)
for r in sm_rows:
    ws_sm.append(r)

ws_sm.append(['psv_nine_toothed_rake', 'HP', 'Percent', '10, 14, 18, 22, 26, 32', desc_str_vn])
ws_sm.append(['psv_nine_toothed_rake', 'DEF', 'Percent', '10, 14, 18, 22, 26, 32', desc_str_vn])
print("Updated StaticModifiers for psv_nine_toothed_rake (HP & DEF)")

# 1.4 Sheet CombatEvents
ws_ce = wb_gc['CombatEvents']
ce_rows = []
for r in list(ws_ce.iter_rows(values_only=True)):
    if r[0] != 'psv_nine_toothed_rake':
        ce_rows.append(r)
ws_ce.delete_rows(1, ws_ce.max_row)
for r in ce_rows:
    ws_ce.append(r)

ws_ce.append(['psv_nine_toothed_rake', 'OnTakeDamage', 'Eff_Stackable_DEF_Buff', '4.0, 5.0, 6.0, 7.0, 8.0, 10.0', 'None', 5, 0, 'self', 'Mỗi khi chịu sát thương, Phòng Thủ tăng thêm {1}%, tối đa cộng dồn 5 tầng. Khi đạt 5 tầng, hồi phục 10% HP Tối Đa.'])
print("Updated CombatEvents for psv_nine_toothed_rake")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)
ws_str = wb_loc['STR']

found = False
for row in ws_str.iter_rows(values_only=False):
    if row[0].value == 'STR_NINE_TOOLTHED_RAKE_SKILL':
        row[1].value = desc_str_en
        row[2].value = desc_str_vn
        found = True
        print("Updated STR_NINE_TOOLTHED_RAKE_SKILL in Localizations.xlsx")
        break

if not found:
    ws_str.append(['STR_NINE_TOOLTHED_RAKE_SKILL', desc_str_en, desc_str_vn])

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
