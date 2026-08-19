import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Tool/data/Localizations.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if row[0] and ('_Des' in str(row[0]) or '_DES' in str(row[0])):
            # print if it looks like a skill (e.g. contains character name or _B_ / _M_ / _U_)
            k = str(row[0])
            if any(x in k for x in ['_B_', '_M_', '_U_', 'SunWukong', 'ErlangShen', 'ZhuBaJie', 'LiJing', 'BullDemonKing', 'ThirdPrinceNezha', 'TangSanZang', 'LittleWhiteDragon', 'ShaWujing']):
                print(f'{k}: {row[2]}')
