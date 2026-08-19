import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Tool/data/Localizations.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if row[0] and 'BullDemonKing' in str(row[0]):
            print(f'[{sheet}] {row[0]}: VI="{row[2]}" | EN="{row[1]}"')
