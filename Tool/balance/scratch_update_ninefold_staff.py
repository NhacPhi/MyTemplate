import openpyxl
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

gc_path = 'Tool/data/GameConfig.xlsx'
loc_path = 'Tool/data/Localizations.xlsx'

ninefold_vi = 'Tăng {0}% Tốc Độ. Tăng thêm {1}% hiệu quả cho các kỹ năng Cường Hóa và Hỗ Trợ.'
ninefold_en = 'Increases SPEED by {0}%. Increases the effectiveness of Enhancement and Support skills by {1}%.'

for attempt in range(5):
    try:
        wb_gc = openpyxl.load_workbook(gc_path)

        # Update Passives
        ws_passives = wb_gc['Passives']
        for row in ws_passives.iter_rows(values_only=False):
            if row[0].value == 'psv_ninefold_staff':
                row[1].value = 'STR_NINEFOLD_STAFF_SKILL'
                row[2].value = ninefold_vi

        # Update StaticModifiers
        ws_static = wb_gc['StaticModifiers']
        for row in ws_static.iter_rows(values_only=False):
            if row[0].value == 'psv_ninefold_staff':
                row[4].value = ninefold_vi

        # Update CombatEvents
        ws_events = wb_gc['CombatEvents']
        for row in ws_events.iter_rows(values_only=False):
            if row[0].value == 'psv_ninefold_staff':
                row[8].value = ninefold_vi

        wb_gc.save(gc_path)
        print("Successfully saved GameConfig.xlsx for Ninefold_Staff!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} failed: {e}. Retrying...")
        time.sleep(1)

for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook(loc_path)
        ws_str = wb_loc['STR']
        for row in ws_str.iter_rows(values_only=False):
            if row[0].value == 'STR_NINEFOLD_STAFF_SKILL':
                row[1].value = ninefold_en
                row[2].value = ninefold_vi

        wb_loc.save(loc_path)
        print("Successfully saved Localizations.xlsx for Ninefold_Staff!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} loc failed: {e}. Retrying...")
        time.sleep(1)
