import openpyxl
import time
import sys

sys.stdout.reconfigure(encoding='utf-8')

gc_path = 'Tool/data/GameConfig.xlsx'
loc_path = 'Tool/data/Localizations.xlsx'

# 1. Moonlit_Firelfy
moonlit_vi = 'Tăng {0}% Tốc Độ và {1}% Tấn Công. Tăng thêm {2}% hiệu quả cho các kỹ năng Cường Hóa và Hỗ Trợ.'
moonlit_en = 'Increases SPEED by {0}% and ATK by {1}%. Increases the effectiveness of Enhancement and Support skills by {2}%.'

# 2. Wings_of_Phoenix
wings_vi = 'Tăng {0}% HP và {1}% Phòng Thủ. Tăng thêm {2}% hiệu quả cho các kỹ năng Hồi Máu.'
wings_en = 'Increases HP by {0}% and DEF by {1}%. Increases the effectiveness of Healing skills by {2}%.'

for attempt in range(5):
    try:
        wb_gc = openpyxl.load_workbook(gc_path)

        # Update Passives
        ws_passives = wb_gc['Passives']
        for row in ws_passives.iter_rows(values_only=False):
            if row[0].value == 'psv_moonlit_firefly':
                row[1].value = 'STR_MOONLIT_FIREFLY_SKILL'
                row[2].value = moonlit_vi
            elif row[0].value == 'psv_wings_of_phoenix':
                row[1].value = 'STR_WINGS_OF_PHOENIX_SKILL'
                row[2].value = wings_vi

        # Update StaticModifiers
        ws_static = wb_gc['StaticModifiers']
        rows_to_keep = []
        for row in ws_static.iter_rows(values_only=True):
            if row[0] not in ['psv_moonlit_firefly', 'psv_wings_of_phoenix']:
                rows_to_keep.append(row)

        ws_static.delete_rows(1, ws_static.max_row)
        for r in rows_to_keep:
            ws_static.append(r)

        # Static for Moonlit Firefly
        ws_static.append(['psv_moonlit_firefly', 'SPEED', 'Percent', '8.0, 10.0, 12.0, 15.0, 18.0, 25.0', moonlit_vi])
        ws_static.append(['psv_moonlit_firefly', 'ATK', 'Percent', '10.0, 12.0, 14.0, 17.0, 20.0, 25.0', moonlit_vi])

        # Static for Wings of Phoenix
        ws_static.append(['psv_wings_of_phoenix', 'HP', 'Percent', '15.0, 18.0, 21.0, 25.0, 30.0, 40.0', wings_vi])
        ws_static.append(['psv_wings_of_phoenix', 'DEF', 'Percent', '10.0, 12.0, 14.0, 17.0, 20.0, 25.0', wings_vi])

        # Update CombatEvents
        ws_events = wb_gc['CombatEvents']
        rows_events_keep = []
        for row in ws_events.iter_rows(values_only=True):
            if row[0] not in ['psv_moonlit_firefly', 'psv_wings_of_phoenix']:
                rows_events_keep.append(row)

        ws_events.delete_rows(1, ws_events.max_row)
        for r in rows_events_keep:
            ws_events.append(r)

        # CombatEvent for Moonlit Firefly
        ws_events.append([
            'psv_moonlit_firefly',
            'OnBeforeDealDamage',
            'Eff_Buff_Enhance',
            '40.0, 50.0, 60.0, 70.0, 80.0, 100.0',
            'None',
            0,
            0,
            'self',
            moonlit_vi
        ])

        # CombatEvent for Wings of Phoenix
        ws_events.append([
            'psv_wings_of_phoenix',
            'OnBeforeDealDamage',
            'Eff_Buff_Enhance',
            '30.0, 40.0, 50.0, 60.0, 70.0, 80.0',
            'None',
            0,
            0,
            'self',
            wings_vi
        ])

        wb_gc.save(gc_path)
        print("Successfully saved GameConfig.xlsx for Moonlit Firefly and Wings of Phoenix!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} failed: {e}. Retrying...")
        time.sleep(1)

for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook(loc_path)
        ws_str = wb_loc['STR']
        for row in ws_str.iter_rows(values_only=False):
            if row[0].value == 'STR_MOONLIT_FIREFLY_SKILL':
                row[1].value = moonlit_en
                row[2].value = moonlit_vi
            elif row[0].value == 'STR_WINGS_OF_PHOENIX_SKILL':
                row[1].value = wings_en
                row[2].value = wings_vi

        wb_loc.save(loc_path)
        print("Successfully saved Localizations.xlsx for Moonlit Firefly and Wings of Phoenix!")
        break
    except Exception as e:
        print(f"Attempt {attempt+1} loc failed: {e}. Retrying...")
        time.sleep(1)
