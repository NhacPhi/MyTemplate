import openpyxl
import sys
import time
sys.stdout.reconfigure(encoding='utf-8')

# 1. Update Localizations.xlsx
for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook('Tool/data/Localizations.xlsx')

        # Update Battle sheet
        ws_battle = wb_loc['Battle']
        battle_updates = {
            'BullDemonKing_B_Name': ('Heavenly Equaling Slash', 'Bình Thiên Trảm'),
            'BullDemonKing_M_Name': ("Demon King's Vigorous Might", 'Ma Vương Bạt Khí'),
            'BullDemonKing_U_Name': ('Bishui Beast Strike', 'Bích Thủy Tinh Thú Kích'),
            'BullDemonKing_B_Des': (
                'Swings a massive blade with earth-shattering force, dealing {0}% ATK damage to a single enemy.',
                'Vung đại đao chém tựa dời núi lấp biển, gây ST bằng {0}% Tấn Công cho một kẻ địch.'
            ),
            'BullDemonKing_M_Des': (
                'The Bull Demon King roars, unleashing surging demonic aura to enter [Heaven-Equaling] state, increasing ATK by {0}% for {1} turns.',
                'Ngưu Ma Vương gầm lên một tiếng làm chấn động càn khôn, bộc phát ma khí ngút trời giúp bản thân vào trạng thái [Bình Thiên], tăng {0}% Tấn Công duy trì {1} hiệp.'
            ),
            'BullDemonKing_U_Des': (
                'The Bishui Golden-Eyed Beast unleashes a shockwave dealing {0}% ATK damage to a single enemy and afflicting them with [Deadly Poison], dealing {1}% ATK Poison damage each turn for {2} turns.',
                'Bích Thủy Kim Tinh Thú phóng ra luồng sóng chấn động gây ST bằng {0}% Tấn Công cho một kẻ địch, đồng thời khiến đối phương rơi vào trạng thái [Kịch Độc], mỗi hiệp chịu ST Độc bằng {1}% Tấn Công trong {2} hiệp.'
            )
        }

        for row in ws_battle.iter_rows(min_row=2):
            key = row[0].value
            if key in battle_updates:
                row[1].value = battle_updates[key][0]
                row[2].value = battle_updates[key][1]

        # Update STR sheet
        ws_str = wb_loc['STR']
        str_keys = {row[0].value for row in ws_str.iter_rows(min_row=2)}
        if 'STR_BULL_COUNTER_NAME' not in str_keys:
            ws_str.append(['STR_BULL_COUNTER_NAME', 'Demon King Retaliation', 'Ma Vương Phản Kích'])
        if 'STR_BULL_COUNTER_DES' not in str_keys:
            ws_str.append(['STR_BULL_COUNTER_DES', 'Upon taking damage, immediately counter-attacks dealing {0}% ATK damage to the attacker.', 'Khi bị kẻ địch tấn công, lập tức phản đòn gây {0}% Tấn Công lên kẻ tấn công.'])

        wb_loc.save('Tool/data/Localizations.xlsx')
        print('Successfully updated Localizations.xlsx')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: Localizations.xlsx locked, retrying in 1s...')
        time.sleep(1)

# 2. Update GameConfig.xlsx
for attempt in range(5):
    try:
        wb_game = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')

        # Update SkillConfig sheet
        ws_skill = wb_game['SkillConfig']
        for row in ws_skill.iter_rows(min_row=2):
            k = row[0].value
            if k == 'BullDemonKing_B':
                row[3].value = 'Melee'
                row[4].value = '1, 1.2, 1.5'
                row[5].value = '0, 0, 0'
                row[6].value = 'BasicAttack'
                row[7].value = 'SingleEnemy'
                row[8].value = 'None'
                row[9].value = 'VanguardTiger_Attack'
                row[10].value = 'psv_bulldemonking_counter'
                print('Updated SkillConfig: BullDemonKing_B')
            elif k == 'BullDemonKing_M':
                row[3].value = 'StatModifier'
                row[4].value = '0.3, 0.4, 0.5'
                row[5].value = '4.0, 4.0, 3.0'
                row[6].value = 'ActiveSkill'
                row[7].value = 'Self'
                row[8].value = 'EFF_Bull_Buff_ATK'
                row[9].value = 'BullDemonKing_Major'
                row[10].value = 'None'
                print('Updated SkillConfig: BullDemonKing_M')
            elif k == 'BullDemonKing_U':
                row[3].value = 'PoisonBall'
                row[4].value = '2.5, 3.0, 3.5'
                row[5].value = '5.0, 5.0, 4.0'
                row[6].value = 'ActiveSkill'
                row[7].value = 'SingleEnemy'
                row[8].value = 'EFF_Bull_Poison'
                row[9].value = 'BullDemonKing_Main'
                row[10].value = 'None'
                print('Updated SkillConfig: BullDemonKing_U')

        # Update EffectConfig sheet
        ws_eff = wb_game['EffectConfig']
        eff_keys = {row[0].value: row for row in ws_eff.iter_rows(min_row=2)}

        if 'EFF_Bull_Buff_ATK' in eff_keys:
            r = eff_keys['EFF_Bull_Buff_ATK']
            r[3].value = 'StatBuff'
            r[4].value = 'ATK'
            r[5].value = 'Percent'
            r[6].value = 3
            r[7].value = 30
            r[8].value = 1
        else:
            ws_eff.append(['EFF_Bull_Buff_ATK', 'BUFF_ATTACK_NAME', 'BUFF_ATTACK_DES', 'StatBuff', 'ATK', 'Percent', 3, 30, 1])

        if 'EFF_Bull_Poison' in eff_keys:
            r = eff_keys['EFF_Bull_Poison']
            r[3].value = 'Poison'
            r[4].value = 'None'
            r[5].value = 'Percent'
            r[6].value = 3
            r[7].value = 8
            r[8].value = 1
        else:
            ws_eff.append(['EFF_Bull_Poison', 'EFF_POSION_NAME', 'EFF_POSION_DES', 'Poison', 'None', 'Percent', 3, 8, 1])

        # Update Passives sheet
        ws_pass = wb_game['Passives']
        pass_keys = {row[0].value: row for row in ws_pass.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in pass_keys:
            ws_pass.append(['psv_bulldemonking_counter', 'STR_BULL_COUNTER_NAME', 'Phản đòn khi bị tấn công gây 60%, 80%, 100% ATK', 'STR_BULL_COUNTER_DES'])
            print('Added psv_bulldemonking_counter to Passives')

        # Update CombatEvents sheet
        ws_ce = wb_game['CombatEvents']
        ce_keys = {row[0].value: row for row in ws_ce.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in ce_keys:
            ws_ce.append(['psv_bulldemonking_counter', 'OnTakeDamage', 'Eff_CounterAttack', 'Always', 'enemy', '60, 80, 100, 100, 100, 100'])
            print('Added psv_bulldemonking_counter to CombatEvents')

        wb_game.save('Tool/data/GameConfig.xlsx')
        print('Successfully updated GameConfig.xlsx')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: GameConfig.xlsx locked, retrying in 1s...')
        time.sleep(1)
