import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 1. Update GameConfig.xlsx
gc_path = 'Tool/data/GameConfig.xlsx'
wb_gc = openpyxl.load_workbook(gc_path)

# 1.1 Sheet EffectConfig
ws_ec = wb_gc['EffectConfig']
eff_found = False
for row in ws_ec.iter_rows(values_only=False):
    if row[0].value == 'EFF_Zhu_DefShred':
        row[1].value = 'DEBUFF_DEF_NAME'
        row[2].value = 'DEBUFF_DEF_DES'
        row[3].value = 'StatDebuff'
        row[4].value = 'DEF'
        row[5].value = 'Percent'
        row[6].value = 2
        row[7].value = -20
        row[8].value = 1
        eff_found = True
        print("Updated EFF_Zhu_DefShred in EffectConfig sheet")
        break

if not eff_found:
    ws_ec.append(['EFF_Zhu_DefShred', 'DEBUFF_DEF_NAME', 'DEBUFF_DEF_DES', 'StatDebuff', 'DEF', 'Percent', 2, -20, 1])
    print("Added EFF_Zhu_DefShred to EffectConfig sheet")

# 1.2 Sheet SkillConfig
ws_sc = wb_gc['SkillConfig']
for row in ws_sc.iter_rows(values_only=False):
    if row[0].value in ['ZhuBaJie_U', 'MarshalTianpeng_U']:
        row[4].value = '1.4, 1.6, 1.8'
        row[8].value = 'EFF_Zhu_DefShred'
        print(f"Updated {row[0].value} Effect_ID to EFF_Zhu_DefShred and Multiplier to '1.4, 1.6, 1.8'")

wb_gc.save(gc_path)
print("Saved GameConfig.xlsx successfully!")

# 2. Update Localizations.xlsx
loc_path = 'Tool/data/Localizations.xlsx'
wb_loc = openpyxl.load_workbook(loc_path)

# 2.1 Sheet Battle
ws_battle = wb_loc['Battle']

vn_desc = "Thúc động bộc phát linh lực hóa thành cự thần khổng lồ, dốc toàn lực giáng một đòn trời giệt đất nát xuống mục tiêu, gây ST bằng {0}% Tấn Công cho một kẻ địch, đồng thời khiến đối phương rơi vào trạng thái [Suy Yếu], giảm 20% Phòng Thủ trong 2 hiệp."
en_desc = "Unleashes divine giant power, slamming down the Nine-Toothed Rake to deal {0}% ATK to an enemy and inflicts [Weakened], reducing target's DEF by 20% for 2 turns."

for row in ws_battle.iter_rows(values_only=False):
    if row[0].value in ['ZhuBaJie_U_Des', 'MarshalTianpeng_U_Des']:
        row[1].value = en_desc
        row[2].value = vn_desc
        print(f"Updated {row[0].value} in Localizations.xlsx")

# 2.2 Sheet Effect (or Battle) for DEBUFF_DEF_NAME / DEBUFF_DEF_DES
ws_eff = wb_loc['Effect'] if 'Effect' in wb_loc.sheetnames else ws_battle
def_name_found = False
def_des_found = False
for row in ws_eff.iter_rows(values_only=False):
    if row[0].value == 'DEBUFF_DEF_NAME':
        row[1].value = 'DEF Reduction'
        row[2].value = 'Giảm Phòng Thủ'
        def_name_found = True
    elif row[0].value == 'DEBUFF_DEF_DES':
        row[1].value = "Reduces target's DEF by {0}%."
        row[2].value = 'Giảm {0}% Phòng Thủ của mục tiêu.'
        def_des_found = True

if not def_name_found:
    ws_eff.append(['DEBUFF_DEF_NAME', 'DEF Reduction', 'Giảm Phòng Thủ'])
if not def_des_found:
    ws_eff.append(['DEBUFF_DEF_DES', "Reduces target's DEF by {0}%.", 'Giảm {0}% Phòng Thủ của mục tiêu.'])

wb_loc.save(loc_path)
print("Saved Localizations.xlsx successfully!")
