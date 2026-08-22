import openpyxl, sys, re

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')
wb_loc = openpyxl.load_workbook('Tool/data/Localizations.xlsx')

loc_dict = {}
for r in wb_loc['STR'].iter_rows(values_only=True):
    if r[0]:
        loc_dict[r[0]] = (r[1], r[2])

passives = {}
for r in wb['Passives'].iter_rows(values_only=True):
    if r[0] and r[0] != 'ID':
        passives[r[0]] = {'key': r[1], 'vn': r[2], 'static': [], 'events': []}

for r in wb['StaticModifiers'].iter_rows(values_only=True):
    if r[0] in passives:
        passives[r[0]]['static'].append((r[1], r[3]))

for r in wb['CombatEvents'].iter_rows(values_only=True):
    if r[0] in passives:
        passives[r[0]]['events'].append((r[1], r[2], r[3]))

print("=== CHECKING FORMAT COMPATIBILITY ===")
for p_id, p in passives.items():
    vn_text = loc_dict.get(p['key'], (None, p['vn']))[1]
    en_text = loc_dict.get(p['key'], (p['vn'], None))[0]
    
    # Collect all args
    args = []
    for s in p['static']:
        vals = [float(x.strip()) for x in str(s[1]).split(',') if x.strip()]
        if vals:
            args.append(vals[0])
    for e in p['events']:
        vals = [float(x.strip()) for x in str(e[2]).split(',') if x.strip()]
        if vals:
            args.append(vals[0])
            
    # Try formatting
    try:
        formatted_vn = vn_text.format(*args) if vn_text else ""
        print(f"[OK] {p_id}: {formatted_vn}")
    except Exception as ex:
        print(f"[ERROR] {p_id} ({p['key']}): {ex}")
        print(f"   VN Text: {vn_text}")
        print(f"   Args count {len(args)}: {args}")
