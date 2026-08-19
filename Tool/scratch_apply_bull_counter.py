import openpyxl
import sys
import time
sys.stdout.reconfigure(encoding='utf-8')

# 1. Update Localizations.xlsx
for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook('Tool/data/Localizations.xlsx')
        ws_battle = wb_loc['Battle']
        for row in ws_battle.iter_rows(min_row=2):
            if row[0].value == 'BullDemonKing_B_Des':
                row[1].value = 'Swings a massive blade dealing {0}% ATK damage to a single enemy. When hit by a basic attack, has a {1}% chance to [Counter-Attack].'
                row[2].value = 'Vung đại đao chém tựa dời núi lấp biển, gây ST bằng {0}% Tấn Công cho một kẻ địch. Khi bị tấn công thường có {1}% cơ hội [Phản Kích].'
                print('Updated BullDemonKing_B_Des in Localizations.xlsx')
        wb_loc.save('Tool/data/Localizations.xlsx')
        print('Successfully saved Localizations.xlsx')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: Localizations.xlsx locked, retrying in 1s...')
        time.sleep(1)

# 2. Update GameConfig.xlsx
for attempt in range(5):
    try:
        wb_game = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')

        # SkillConfig
        ws_skill = wb_game['SkillConfig']
        for row in ws_skill.iter_rows(min_row=2):
            if row[0].value == 'BullDemonKing_B':
                row[3].value = 'Melee'
                row[4].value = '1, 1.2, 1.5'
                row[5].value = '0, 0, 0'
                row[6].value = 'BasicAttack'
                row[7].value = 'SingleEnemy'
                row[8].value = 'None'
                row[9].value = 'VanguardTiger_Attack'
                row[10].value = 'psv_bulldemonking_counter'
                print('Updated BullDemonKing_B in SkillConfig')

        # Passives
        ws_pass = wb_game['Passives']
        pass_keys = {row[0].value: row for row in ws_pass.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in pass_keys:
            ws_pass.append(['psv_bulldemonking_counter', 'STR_BULL_COUNTER_NAME', 'Có 50%, 75%, 100% cơ hội phản kích khi bị đánh thường', 'STR_BULL_COUNTER_DES'])
            print('Added psv_bulldemonking_counter to Passives')
        else:
            r = pass_keys['psv_bulldemonking_counter']
            r[2].value = 'Có 50%, 75%, 100% cơ hội phản kích khi bị đánh thường'

        # CombatEvents
        ws_ce = wb_game['CombatEvents']
        ce_keys = {row[0].value: row for row in ws_ce.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' not in ce_keys:
            ws_ce.append(['psv_bulldemonking_counter', 'OnTakeDamage', 'Eff_CounterAttack', 'BasicAttack', 'enemy', '50, 75, 100, 100, 100, 100'])
            print('Added psv_bulldemonking_counter to CombatEvents')
        else:
            r = ce_keys['psv_bulldemonking_counter']
            r[1].value = 'OnTakeDamage'
            r[2].value = 'Eff_CounterAttack'
            r[3].value = 'BasicAttack'
            r[4].value = 'enemy'
            r[5].value = '50, 75, 100, 100, 100, 100'

        wb_game.save('Tool/data/GameConfig.xlsx')
        print('Successfully saved GameConfig.xlsx')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: GameConfig.xlsx locked, retrying in 1s...')
        time.sleep(1)
