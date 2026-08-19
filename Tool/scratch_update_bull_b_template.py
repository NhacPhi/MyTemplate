import openpyxl
import sys
import time
sys.stdout.reconfigure(encoding='utf-8')

# 1. Update Localizations.xlsx
for attempt in range(5):
    try:
        wb_loc = openpyxl.load_workbook('Tool/data/Localizations.xlsx')
        ws_battle = wb_loc['Battle']
        for row in ws_battle.iter_rows(min_row=2):
            if row[0].value == 'BullDemonKing_B_Des':
                row[1].value = 'Swings a massive blade dealing {0}% ATK damage to a single enemy. \\nWhen hit by a basic attack, has a {1}% chance to [Counter-Attack], dealing {2}% ATK damage.'
                row[2].value = 'Vung đại đao chém tựa dời núi lấp biển, gây ST bằng {0}% Tấn Công cho một kẻ địch. \\nKhi bị tấn công thường có {1}% cơ hội [Phản Kích], đòn đánh gây ST {2}% Tấn Công.'
                print('Updated BullDemonKing_B_Des in Localizations.xlsx')
        wb_loc.save('Tool/data/Localizations.xlsx')
        print('Successfully saved Localizations.xlsx')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: Localizations.xlsx locked, retrying in 1s...')
        time.sleep(1)

# 2. Update GameConfig.xlsx
for attempt in range(5):
    try:
        wb_game = openpyxl.load_workbook('Tool/data/GameConfig.xlsx')

        # CombatEvents
        ws_ce = wb_game['CombatEvents']
        ce_keys = {row[0].value: row for row in ws_ce.iter_rows(min_row=2)}
        if 'psv_bulldemonking_counter' in ce_keys:
            r = ce_keys['psv_bulldemonking_counter']
            r[1].value = 'OnTakeDamage'
            r[2].value = 'Eff_CounterAttack'
            r[3].value = '50, 75, 100, 100, 100, 100'
            r[4].value = 'BasicAttack'
            r[5].value = 100 # effect_param = 100% ATK counter damage
            r[6].value = 0
            r[7].value = 'enemy'
            r[8].value = 'Có 50%, 75%, 100% cơ hội phản kích khi bị đánh thường, gây 100% ATK'
            print('Updated psv_bulldemonking_counter in CombatEvents')

        wb_game.save('Tool/data/GameConfig.xlsx')
        print('Successfully saved GameConfig.xlsx!')
        break
    except PermissionError:
        print(f'Attempt {attempt+1}: GameConfig.xlsx locked, retrying in 1s...')
        time.sleep(1)
